from __future__ import annotations

from io import BytesIO
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def build_workbook(
    *,
    report_name: str,
    generated_at: str,
    filters: Mapping[str, Any] | None,
    summary: Mapping[str, Any] | None,
    rows: Sequence[Mapping[str, Any]],
    column_order: Sequence[str],
    column_titles: Mapping[str, str] | None = None,
) -> bytes:
    """
    Build a simple XLSX with:
      - Data sheet: table rows with deterministic column order (first/active)
      - Summary sheet: report metadata, filters, summary
    Returns XLSX bytes.
    """
    wb = Workbook()

    bold = Font(bold=True)
    # Data sheet first (Excel opens this by default).
    ws_data = wb.active
    ws_data.title = "Data"

    titles = column_titles or {}
    header = [titles.get(k, k) for k in column_order]
    ws_data.append(header)
    for cell in ws_data[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws_data.append([row.get(k) for k in column_order])

    ws_data.freeze_panes = "A2"
    ws_data.auto_filter.ref = ws_data.dimensions

    # Column sizing based on title and a small sample of rows.
    sample_rows = min(100, len(rows))
    for idx, key in enumerate(column_order, start=1):
        title = str(titles.get(key, key))
        max_len = len(title)
        for i in range(sample_rows):
            val = rows[i].get(key) if i < len(rows) else None
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        width = max(12, min(52, max_len + 2))
        ws_data.column_dimensions[ws_data.cell(row=1, column=idx).column_letter].width = width

    # Summary sheet second.
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Report Name"
    ws_summary["B1"] = report_name
    ws_summary["A2"] = "Generated At"
    ws_summary["B2"] = generated_at

    ws_summary["A1"].font = bold
    ws_summary["A2"].font = bold

    r = 4
    ws_summary[f"A{r}"] = "Filters"
    ws_summary[f"A{r}"].font = bold
    r += 1
    if filters:
        for k in sorted(filters.keys()):
            ws_summary[f"A{r}"] = str(k)
            ws_summary[f"B{r}"] = "" if filters[k] is None else str(filters[k])
            r += 1
    else:
        ws_summary[f"A{r}"] = "(none)"
        r += 1

    r += 1
    ws_summary[f"A{r}"] = "Summary"
    ws_summary[f"A{r}"].font = bold
    r += 1
    if summary:
        for k in sorted(summary.keys()):
            ws_summary[f"A{r}"] = str(k)
            ws_summary[f"B{r}"] = "" if summary[k] is None else str(summary[k])
            r += 1
    else:
        ws_summary[f"A{r}"] = "(none)"
        r += 1

    for col in ("A", "B"):
        ws_summary.column_dimensions[col].width = 28

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
